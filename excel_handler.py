# excel_handler.py
"""
============================================================================
РАБОТА С EXCEL
============================================================================
"""

import pandas as pd
import re
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import COUNTRY_SCHEMA_MAP, PASSPORT_VALIDITY, ITR_KEYWORDS, ALL_COLUMNS

__all__ = [
    "ALL_COLUMNS",
    "save_as_excel",
    "create_application_row",
    "create_empty_row",
    "format_date_ddmmyyyy",
    "merge_employee_into_application_row",
]

IULIIA_AVAILABLE = False
try:
    import iuliia
    IULIIA_AVAILABLE = True
except ImportError:
    pass


def format_date_ddmmyyyy(date_val) -> str:
    if not date_val or str(date_val) == "nan" or str(date_val).strip() == "":
        return ""

    date_str = str(date_val).strip()

    if re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_str):
        return date_str

    formats = [
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%d.%m.%Y")
        except:
            continue

    match = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{day.zfill(2)}.{month.zfill(2)}.{year}"

    return date_str


def transliterate_name(name: str, citizenship: str = None) -> str:
    if not name:
        return ""

    schema = "gost_52535"
    if citizenship:
        for country, s in COUNTRY_SCHEMA_MAP.items():
            if country in citizenship.upper():
                schema = s
                break

    if IULIIA_AVAILABLE:
        try:
            return iuliia.translate(name, schema=iuliia.Schemas.get(schema)).upper()
        except:
            pass

    table = {
        'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'E',
        'Ж':'ZH','З':'Z','И':'I','Й':'Y','К':'K','Л':'L','М':'M',
        'Н':'N','О':'O','П':'P','Р':'R','С':'S','Т':'T','У':'U',
        'Ф':'F','Х':'KH','Ц':'TS','Ч':'CH','Ш':'SH','Щ':'SHCH',
        'Ы':'Y','Э':'E','Ю':'YU','Я':'YA','Ь':'','Ъ':'',
    }
    return ''.join(table.get(c, c) for c in str(name).upper())


def calculate_passport_expiry(doc_date: str, citizenship: str) -> str:
    """Расчет срока действия паспорта - возвращает дату в формате ДД.ММ.ГГГГ"""
    if not doc_date or str(doc_date) == "nan":
        return ""

    cu = str(citizenship).upper().strip() if citizenship else ""

    if cu in ["РОССИЯ", "РФ"]:
        return ""

    if cu == "БЕЛАРУСЬ":
        return "проверить паспорт"

    if cu in ["КИРГИЗИЯ", "СЕРБИЯ"]:
        try:
            from datetime import timedelta
            for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    dt = datetime.strptime(str(doc_date), fmt)
                    break
                except:
                    continue
            else:
                match = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', str(doc_date))
                if match:
                    day, month, year = match.groups()
                    dt = datetime(int(year), int(month), int(day))
                else:
                    return ""

            expiry = dt + timedelta(days=120*30)
            return expiry.strftime("%d.%m.%Y")
        except:
            return ""

    if cu in PASSPORT_VALIDITY:
        try:
            from datetime import timedelta
            for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    dt = datetime.strptime(str(doc_date), fmt)
                    break
                except:
                    continue
            else:
                match = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', str(doc_date))
                if match:
                    day, month, year = match.groups()
                    dt = datetime(int(year), int(month), int(day))
                else:
                    return ""

            expiry = dt + timedelta(days=120*30 - 1)
            return expiry.strftime("%d.%m.%Y")
        except:
            return ""

    return ""


def get_classification(position: str) -> str:
    if not position:
        return "Рабочие"
    pos_lower = position.lower()
    for keyword in ITR_KEYWORDS:
        if keyword in pos_lower:
            return "ИТР"
    return "Рабочие"


def get_department_category(department: str) -> str:
    """Определение категории отдела из названия подразделения"""
    if not department:
        return ""
    dept_upper = department.upper()
    if "СМУ" in dept_upper:
        return "СМУ"
    elif "УМИТ" in dept_upper or "УМиТ" in dept_upper:
        return "УМиТ"
    elif "ОТИЗ" in dept_upper or "ОТ и З" in dept_upper:
        return "ОТиЗ"
    elif "ПТО" in dept_upper:
        return "ПТО"
    elif "ОКС" in dept_upper:
        return "ОКС"
    elif "СДО" in dept_upper:
        return "СДО"
    elif "ОХРАН" in dept_upper:
        return "Охрана"
    elif "СКЛАД" in dept_upper:
        return "Склад"
    elif "ТРАНСП" in dept_upper:
        return "Транспорт"
    elif "МЕХАН" in dept_upper:
        return "Механизация"
    elif "ЭЛЕКТР" in dept_upper:
        return "Электро"
    elif "ЛАБОРАТ" in dept_upper:
        return "Лаборатория"
    elif "КАДР" in dept_upper:
        return "Управление"
    return ""


def get_document_type(citizenship: str, doc_series: str) -> str:
    cu = str(citizenship).upper().strip() if citizenship else ""

    if "РОССИЯ" in cu or cu == "РФ":
        return "Паспорт гражданина России"

    series_clean = str(doc_series).strip()
    if series_clean in ["82", "83"]:
        return "Вид на жительство"

    return "Паспорт иностранного гражданина"


def resolve_row_department(employee: Dict = None, login_department: str = "") -> str:
    """Подразделение в заявке = из базы сотрудника (столбец Подразделение), иначе логин ОП."""
    if employee:
        dept = str(employee.get("department", "") or "").strip()
        if dept:
            return dept
    return login_department or ""


def employee_to_application_row(employee: Dict, row_num: int = 1) -> Dict:
    """Строка заявки из записи базы."""
    pdf_data = {
        "fio": employee.get("fio", ""),
        "route": "",
        "date": "",
        "reason": "",
        "phone": employee.get("phone", ""),
    }
    dept = resolve_row_department(employee, "")
    return create_application_row(row_num, dept, pdf_data, employee)


def merge_employee_into_application_row(
    existing_row: Dict,
    employee: Dict,
    row_num: int = None,
) -> Dict:
    """
    Подставить данные сотрудника из базы, сохранив распознанные поля заявки:
    маршрут, дата вылета, обоснование, АВИА/ЖД, примечание (кроме метки «не найден»).
    """
    row_num = row_num or existing_row.get("№", 1)
    dept = existing_row.get("Подразделение", "") or resolve_row_department(employee, "")
    note = str(existing_row.get("Примечание", "") or "").strip()
    for marker in ("НЕ НАЙДЕН В БАЗЕ (ручной ввод)", "НЕ НАЙДЕН В БАЗЕ"):
        note = note.replace(marker, "").strip()

    pdf_data = {
        "fio": existing_row.get("Ф.И.О.", "") or employee.get("fio", ""),
        "route": existing_row.get("Маршрут", ""),
        "date": existing_row.get("Дата вылета", ""),
        "reason": existing_row.get("Обоснование", ""),
        "phone": existing_row.get("Телефон", "") or employee.get("phone", ""),
    }
    row = create_application_row(row_num, dept, pdf_data, employee)
    if existing_row.get("АВИА/ЖД"):
        row["АВИА/ЖД"] = existing_row.get("АВИА/ЖД")
    row["Примечание"] = note
    return row


def create_application_row(row_num: int, department: str, pdf_data: Dict,
                          employee: Dict = None) -> Dict:
    """Создание строки заявки - ДОБАВЛЕНИЕ, а не замена"""
    from database import format_passport_field

    row = {}

    # A - №
    row["№"] = row_num

    # B - Подразделение — из базы сотрудника
    row["Подразделение"] = resolve_row_department(employee, department)

    # === ИСПРАВЛЕНО: C - Отдел = из Базы столбец C - department_category ===
    if employee:
        # Сначала проверяем department_category из базы (столбец C)
        dept_cat = employee.get('department_category', '')
        if not dept_cat:
            # Если нет, берем из department
            dept_cat = get_department_category(employee.get('department', ''))
        row["Отдел"] = dept_cat
    else:
        row["Отдел"] = ""

    # D - Операция
    row["Операция"] = "Заказ"

    # E - Классификация (из Базы)
    row["Классификация"] = get_classification(employee.get('position', '')) if employee else "Рабочие"

    # F - Дата заказа - ДД.ММ.ГГГГ
    row["Дата заказа"] = datetime.now().strftime("%d.%m.%Y")

    # G - Организация
    row["Организация"] = "Монтаж"

    # H - Ф.И.О. (из PDF)
    row["Ф.И.О."] = pdf_data.get('fio', '')

    # I - Ф.И.О лат
    citizenship = employee.get('citizenship', '') if employee else ''
    row["Ф.И.О лат"] = transliterate_name(pdf_data.get('fio', ''), citizenship)

    # J - Табельный номер (из Базы)
    row["Табельный номер"] = employee.get('tab_num', '') if employee else ''

    # K - Гражданство (из Базы)
    row["Гражданство"] = citizenship

    # L - Дата рождения - ДД.ММ.ГГГГ (из Базы)
    row["Дата рождения"] = format_date_ddmmyyyy(employee.get('birth_date', '')) if employee else ''

    # M - Вид документа
    doc_series = employee.get('doc_series', '') if employee else ''
    row["Вид документа"] = get_document_type(citizenship, doc_series)
    row["Серия"] = format_passport_field(doc_series)

    # O - Номер (из Базы)
    row["Номер"] = format_passport_field(employee.get("doc_num", "")) if employee else ""

    # P - Дата выдачи - ДД.ММ.ГГГГ (из Базы)
    row["Дата выдачи"] = format_date_ddmmyyyy(employee.get('doc_date', '')) if employee else ''

    # Q - Дата окончания - ДД.ММ.ГГГГ (ИЗ БАЗЫ - столбец Q)
    if employee:
        # Сначала берем из doc_expiry (столбец Q базы)
        doc_expiry = employee.get('doc_expiry', '')
        if doc_expiry:
            row["Дата окончания"] = format_date_ddmmyyyy(doc_expiry)
        else:
            # Если нет, рассчитываем
            doc_date = employee.get('doc_date', '')
            row["Дата окончания"] = calculate_passport_expiry(doc_date, citizenship)
    else:
        row["Дата окончания"] = ""

    # R - Кем выдан (из Базы)
    row["Кем выдан"] = employee.get('doc_issuer', '') if employee else ''

    # === ИСПРАВЛЕНО: S - Адрес = из Базы столбец S - Место постоянного проживания ===
    row["Адрес"] = employee.get('address', '') if employee else ''

    # T - Маршрут (из PDF)
    row["Маршрут"] = pdf_data.get('route', '')

    # U - Обоснование (из PDF)
    row["Обоснование"] = pdf_data.get('reason', '')

    # V - ПС
    row["ПС"] = ""

    # W - АВИА/ЖД
    row["АВИА/ЖД"] = "АВИА"

    # X - Дата вылета - ДД.ММ.ГГГГ (из PDF)
    row["Дата вылета"] = format_date_ddmmyyyy(pdf_data.get('date', ''))

    # Y - Примечание
    row["Примечание"] = ""

    # Z - Ответственный
    row["Ответственный"] = ""

    # AA - Дата выписки
    row["Дата выписки"] = ""

    # AB - Билет
    row["Билет"] = ""

    # AC - Сумма
    row["Сумма"] = ""

    # AD - Оплата
    row["Оплата"] = "Монтаж"

    # AE - Причина возврата
    row["Причина возврата"] = ""

    # AF - Последний перелет
    row["Последний перелет"] = ""

    # AG - Телефон
    phone = pdf_data.get('phone', '')
    if not phone and employee:
        phone = employee.get('phone', '')
    row["Телефон"] = phone

    # AH - Трансфер
    row["Трансфер"] = ""

    return row


def create_empty_row(row_num: int, department: str, pdf_data: Dict) -> Dict:
    row = {}
    row["№"] = row_num
    row["Подразделение"] = resolve_row_department(None, department)
    row["Отдел"] = ""
    row["Операция"] = "Заказ"
    row["Классификация"] = "Рабочие"
    row["Дата заказа"] = datetime.now().strftime("%d.%m.%Y")
    row["Организация"] = "Монтаж"
    row["Ф.И.О."] = pdf_data.get('fio', '')
    row["Ф.И.О лат"] = ""
    row["Табельный номер"] = ""
    row["Гражданство"] = ""
    row["Дата рождения"] = ""
    row["Вид документа"] = ""
    row["Серия"] = ""
    row["Номер"] = ""
    row["Дата выдачи"] = ""
    row["Дата окончания"] = ""
    row["Кем выдан"] = ""
    row["Адрес"] = ""
    row["Маршрут"] = pdf_data.get('route', '')
    row["Обоснование"] = pdf_data.get('reason', '')
    row["ПС"] = ""
    row["АВИА/ЖД"] = "АВИА"
    row["Дата вылета"] = format_date_ddmmyyyy(pdf_data.get('date', ''))
    row["Примечание"] = "НЕ НАЙДЕН В БАЗЕ"
    row["Ответственный"] = ""
    row["Дата выписки"] = ""
    row["Билет"] = ""
    row["Сумма"] = ""
    row["Оплата"] = "Монтаж"
    row["Причина возврата"] = ""
    row["Последний перелет"] = ""
    row["Телефон"] = pdf_data.get('phone', '')
    row["Трансфер"] = ""
    return row


def save_as_excel(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"

    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)
