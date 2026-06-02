# template_manager.py
"""Шаблон заявки Excel: загрузка, списки НАСТРОЙКА, экспорт с валидацией."""

import pickle
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import ALL_COLUMNS, DATA_DIR
from app_logger import log

TEMPLATE_FILE = DATA_DIR / "application_template.xlsx"
TEMPLATE_META_FILE = DATA_DIR / "template_meta.pkl"
WORK_SHEET_NAMES = ("Заявка", "Sheet1", "ЗАЯВКА")
NASTROYKA_SHEET = "НАСТРОЙКА"


def _read_column_values(ws: Worksheet, col: str, start: int, end: int) -> List[str]:
    values = []
    for row in range(start, end + 1):
        val = ws[f"{col}{row}"].value
        if val is not None and str(val).strip():
            values.append(str(val).strip())
    return values


def parse_nastroyka(wb) -> Dict[str, List[str]]:
    if NASTROYKA_SHEET not in wb.sheetnames:
        return {}
    ws = wb[NASTROYKA_SHEET]
    return {
        "Подразделение": _read_column_values(ws, "A", 2, 32),
        "Отдел": _read_column_values(ws, "B", 2, 20),
        "Операция": _read_column_values(ws, "C", 2, 4),
        "Классификация": _read_column_values(ws, "D", 2, 3),
        "Организация": _read_column_values(ws, "E", 2, 3),
        "Обоснование": _read_column_values(ws, "H", 2, 10),
        "АВИА/ЖД": _read_column_values(ws, "I", 2, 3),
        "Оплата": _read_column_values(ws, "J", 2, 3),
        "Маршрут": _read_column_values(ws, "K", 2, 2122),
        "Ответственный": _read_column_values(ws, "N", 2, 66),
    }


def install_template(source_path: str) -> Tuple[bool, str]:
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, TEMPLATE_FILE)
        wb = load_workbook(TEMPLATE_FILE, data_only=False)
        lists = parse_nastroyka(wb)
        wb.close()
        meta = {"lists": lists, "source": source_path, "installed": datetime.now().isoformat()}
        with open(TEMPLATE_META_FILE, "wb") as f:
            pickle.dump(meta, f)
        log.info("Шаблон установлен: %s", source_path)
        return True, f"Шаблон загружен. Списков НАСТРОЙКА: {len(lists)}"
    except Exception as e:
        log.error("Шаблон: %s", e)
        return False, str(e)


def get_dropdown_lists() -> Dict[str, List[str]]:
    if TEMPLATE_META_FILE.exists():
        try:
            # Проверка целостности pickle файла шаблона
            file_size = TEMPLATE_META_FILE.stat().st_size
            if file_size == 0:
                return {}
            if file_size > 10 * 1024 * 1024:  # 10 MB лимит для метаданных шаблона
                log.warning("Файл метаданных шаблона слишком большой")
                return {}
            
            with open(TEMPLATE_META_FILE, "rb") as f:
                meta = pickle.load(f)
            return meta.get("lists", {})
        except Exception:
            pass
    return {}


def is_template_installed() -> bool:
    return TEMPLATE_FILE.exists()


def get_column_dropdowns() -> Dict[str, List[str]]:
    """Сопоставление колонок заявки со списками."""
    lists = get_dropdown_lists()
    return {
        "Подразделение": lists.get("Подразделение", []),
        "Отдел": lists.get("Отдел", []),
        "Операция": lists.get("Операция", ["Заказ"]),
        "Классификация": lists.get("Классификация", []),
        "Организация": lists.get("Организация", ["Монтаж"]),
        "Маршрут": lists.get("Маршрут", []),
        "Обоснование": lists.get("Обоснование", []),
        "АВИА/ЖД": lists.get("АВИА/ЖД", ["АВИА"]),
        "Оплата": lists.get("Оплата", ["Монтаж"]),
        "Ответственный": lists.get("Ответственный", []),
    }


def _find_work_sheet(wb):
    for name in WORK_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _copy_row_validations(ws: Worksheet, from_row: int, to_row: int) -> None:
    if not ws.data_validations:
        return
    for dv in ws.data_validations.dataValidation:
        sqref = str(dv.sqref)
        if f"{from_row}" in sqref:
            new_dv = copy(dv)
            new_sqref = sqref.replace(str(from_row), str(to_row))
            new_dv.sqref = new_sqref
            ws.add_data_validation(new_dv)


def build_export_filename(df: pd.DataFrame, department: str) -> str:
    dep = department.replace("Admin", "Общее").strip()
    dates = []
    if df is not None and not df.empty and "Дата вылета" in df.columns:
        for val in df["Дата вылета"]:
            s = str(val).strip()
            if re.match(r"\d{2}\.\d{2}\.\d{4}", s):
                try:
                    dates.append(datetime.strptime(s, "%d.%m.%Y"))
                except ValueError:
                    pass
    if dates:
        nearest = min(dates, key=lambda d: abs((d - datetime.now()).days))
        date_str = nearest.strftime("%d.%m.%Y")
    else:
        date_str = datetime.now().strftime("%d.%m.%Y")

    count = len(df) if df is not None else 0
    surname = ""
    if df is not None and not df.empty and "Ф.И.О." in df.columns:
        first_fio = str(df.iloc[0]["Ф.И.О."]).strip()
        if first_fio:
            surname = first_fio.split()[0]

    name = f'ОП "{dep}" Заявка на покупку билетов "{date_str}", "{count}" чел.'
    if surname:
        name += f' ("{surname}")'
    return name + ".xlsx"


def export_application(df: pd.DataFrame, output_path: str) -> Tuple[bool, str]:
    if not TEMPLATE_FILE.exists():
        from excel_handler import save_as_excel
        save_as_excel(df, output_path)
        return True, "Экспорт без шаблона (простой Excel)"

    try:
        shutil.copy2(TEMPLATE_FILE, output_path)
        wb = load_workbook(output_path)
        ws = _find_work_sheet(wb)
        start_row = 2
        template_row = 2

        if ws.max_row > start_row:
            ws.delete_rows(start_row + 1, ws.max_row - start_row)

        for i, row_data in enumerate(df.values):
            row_idx = start_row + i
            if row_idx > template_row:
                for col_idx in range(1, len(ALL_COLUMNS) + 1):
                    src = ws.cell(template_row, col_idx)
                    dst = ws.cell(row_idx, col_idx)
                    dst._style = copy(src._style)
                    dst.number_format = src.number_format
                _copy_row_validations(ws, template_row, row_idx)

            for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
                val = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ""
                if pd.isna(val):
                    val = ""
                ws.cell(row_idx, col_idx, value=val)

        wb.save(output_path)
        wb.close()
        return True, f"Файл сохранён: {output_path}"
    except Exception as e:
        log.error("Экспорт шаблона: %s", e)
        return False, str(e)
